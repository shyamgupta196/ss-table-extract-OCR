import os
import argparse
import cv2
from paddleocr import PPStructure, draw_structure_result, save_structure_res
import openpyxl
import numpy as np
import pandas as pd

parser = argparse.ArgumentParser(
    description="Detect table and using OCR store it in excel file."
)


parser.add_argument(
    "--ss_folder",
    default="screenshot",
    type=str,
    required=False,
    help="path to input folder",
)

parser.add_argument(
    "--ss_output_txt",
    default="tables",
    type=str,
    required=False,
    help="path to input folder",
)


parser.add_argument(
    "--input_folder",
    default="extracted_tables",
    type=str,
    required=False,
    help="path to input folder",
)
parser.add_argument(
    "--save_folder",
    default=r".\output_excels",
    type=str,
    required=False,
    help="path to save folder for excel and bboxes",
)

args = parser.parse_args()


def detect_table():
    os.system(f"python recognition.py --input_folder {args.ss_folder} --output_folder {args.ss_output_txt}")


def TableOCR():
    for filename in os.listdir(args.input_folder):
        try:
            if filename.endswith((".jpg", ".png", ".jpeg")):
                path = os.path.join(args.input_folder, filename)
                img = cv2.imread(path)
                result = table_engine(img)
                save_structure_res(
                    result, args.save_folder, os.path.basename(path).split(".")[0]
                )
                os.makedirs(
                    os.path.join(
                        args.save_folder, os.path.basename(path).split(".")[0]
                    ),
                    exist_ok=True,
                )
                try:
                    pathhtml = os.path.join(
                        args.save_folder,
                        filename.split(".")[0],
                        filename.split(".")[0] + ".html",
                    )
                    pathexcel = os.path.join(
                        args.save_folder,
                        filename.split(".")[0],
                        filename.split(".")[0] + ".xlsx",
                    )
                    # import IPython; IPython.embed();exit()
                    html_content = result[0]["res"]["html"]
                    df = pd.read_html(html_content)
                    df = df[0]
                    df.columns = df.iloc[0]
                    df = df[1:]
                    df.to_html(pathhtml)
                    df.to_excel(pathexcel)
                except Exception as e:
                    print(e)
                    workbook = openpyxl.Workbook()
                    worksheet = workbook.active

                    # Add headers for the columns
                    worksheet.cell(row=1, column=1, value="Text")
                    worksheet.cell(row=1, column=2, value="Confidence")
                    worksheet.cell(row=1, column=3, value="x_min")
                    worksheet.cell(row=1, column=4, value="y_min")
                    worksheet.cell(row=1, column=5, value="x_max")
                    worksheet.cell(row=1, column=6, value="y_max")

                    # Iterate through the dictionary and insert data into the worksheet
                    for idx, data in enumerate(result[0]["res"], start=2):
                        text = data["text"]
                        confidence = data["confidence"]
                        text_region = data["text_region"]

                        x_min = min(point[0] for point in text_region)
                        y_min = min(point[1] for point in text_region)
                        x_max = max(point[0] for point in text_region)
                        y_max = max(point[1] for point in text_region)

                        worksheet.cell(row=idx, column=1, value=text)
                        worksheet.cell(row=idx, column=2, value=confidence)
                        worksheet.cell(row=idx, column=3, value=x_min)
                        worksheet.cell(row=idx, column=4, value=y_min)
                        worksheet.cell(row=idx, column=5, value=x_max)
                        worksheet.cell(row=idx, column=6, value=y_max)

                    df = pd.DataFrame(workbook.active.values)
                    df.columns = df.iloc[0]
                    df = df[1:]
                    differences = np.diff(df["x_max"])
                    indices = np.where(-differences > 500)[0]
                    print(df.iloc[indices, :])
                    df["Text"].iloc[indices + 1]
                    data = df["Text"].values
                    subarrays = np.split(data, indices + 1)
                    new_df = pd.DataFrame(subarrays)
                    new_df.columns = new_df.iloc[0]
                    new_df = new_df[1:]
                    new_df.to_excel(
                        os.path.join(
                            args.save_folder,
                            os.path.basename(path).split(".")[0],
                            filename.split(".")[0] + ".xlsx",
                        )
                    )
                    new_df.to_html(
                        os.path.join(
                            args.save_folder,
                            os.path.basename(path).split(".")[0],
                            filename.split(".")[0] + ".html",
                        )
                    )
                    continue
        except Exception as e:
            print(e)
            continue


if __name__ == "__main__":
    table_engine = PPStructure(
        show_log=True, det_db_box_thresh=0.3, det_pse_box_thresh=0.3
    )
    detect_table()
    TableOCR()

## psuedocode
# read output.xlsx
# take x_max
# In [7]: differences = np.diff(df['x_max'])
# In [10]: indices = np.where(-differences > 500)[0]
# In [11]: indices
# df.iloc[indices,:]
# In [17]: df['Text']
# In [19]: df['Text'].iloc[indices+1]
# In [20]: df['Text'].iloc[indices]
# In [22]: data = df['x_max'].values
# In [30]: subarrays = np.split(data, indices + 1)
# In [32]: pd.DataFrame(subarrays)
