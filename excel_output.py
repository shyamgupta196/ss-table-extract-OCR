import openpyxl

result = [
    {
        "text": "Item",
        "confidence": 0.9473044276237488,
        "text_region": [[32.0, 5.0], [66.0, 5.0], [66.0, 24.0], [32.0, 24.0]],
    },
    {
        "text": "Amount",
        "confidence": 0.9938108325004578,
        "text_region": [[99.0, 5.0], [150.0, 5.0], [150.0, 24.0], [99.0, 24.0]],
    },
    {
        "text": "Quantity",
        "confidence": 0.9783017635345459,
        "text_region": [[265.0, 3.0], [319.0, 6.0], [318.0, 26.0], [264.0, 22.0]],
    },
    {
        "text": "Or.",
        "confidence": 0.6895613074302673,
        "text_region": [[436.0, 2.0], [494.0, 2.0], [494.0, 25.0], [436.0, 25.0]],
    },
    {
        "text": "Purchase order",
        "confidence": 0.965871274471283,
        "text_region": [[509.0, 6.0], [608.0, 6.0], [608.0, 24.0], [509.0, 24.0]],
    },
    {
        "text": "Item",
        "confidence": 0.9121732711791992,
        "text_region": [[603.0, 6.0], [648.0, 6.0], [648.0, 24.0], [603.0, 24.0]],
    },
    {
        "text": "POText",
        "confidence": 0.9683301448822021,
        "text_region": [[675.0, 5.0], [729.0, 5.0], [729.0, 24.0], [675.0, 24.0]],
    },
    {
        "text": "TaxCode",
        "confidence": 0.9939543604850769,
        "text_region": [[1078.0, 6.0], [1140.0, 6.0], [1140.0, 24.0], [1078.0, 24.0]],
    },
    {
        "text": "No...A",
        "confidence": 0.903913676738739,
        "text_region": [[1162.0, 5.0], [1213.0, 5.0], [1213.0, 23.0], [1162.0, 23.0]],
    },
    {
        "text": ">",
        "confidence": 0.28845545649528503,
        "text_region": [[5.0, 34.0], [23.0, 34.0], [23.0, 52.0], [5.0, 52.0]],
    },
    {
        "text": "1",
        "confidence": 0.9950351715087891,
        "text_region": [[83.0, 36.0], [94.0, 36.0], [94.0, 50.0], [83.0, 50.0]],
    },
    {
        "text": "1,409,869.30",
        "confidence": 0.9679698944091797,
        "text_region": [[160.0, 33.0], [262.0, 33.0], [262.0, 51.0], [160.0, 51.0]],
    },
    {
        "text": "14,000EA",
        "confidence": 0.9731037616729736,
        "text_region": [[386.0, 33.0], [457.0, 33.0], [457.0, 51.0], [386.0, 51.0]],
    },
    {
        "text": "口",
        "confidence": 0.21123601496219635,
        "text_region": [[483.0, 36.0], [497.0, 36.0], [497.0, 48.0], [483.0, 48.0]],
    },
    {
        "text": "4200001437",
        "confidence": 0.9975188374519348,
        "text_region": [[506.0, 34.0], [592.0, 34.0], [592.0, 52.0], [506.0, 52.0]],
    },
    {
        "text": "140",
        "confidence": 0.9987015724182129,
        "text_region": [[612.0, 32.0], [644.0, 32.0], [644.0, 51.0], [612.0, 51.0]],
    },
    {
        "text": "IN Basiccustoms",
        "confidence": 0.9705668091773987,
        "text_region": [[669.0, 32.0], [777.0, 34.0], [777.0, 52.0], [668.0, 50.0]],
    },
    {
        "text": "GI(GST",
        "confidence": 0.9181970953941345,
        "text_region": [[1075.0, 33.0], [1162.0, 33.0], [1162.0, 51.0], [1075.0, 51.0]],
    },
    {
        "text": "口",
        "confidence": 0.5829343795776367,
        "text_region": [[1173.0, 36.0], [1186.0, 36.0], [1186.0, 48.0], [1173.0, 48.0]],
    },
    {
        "text": "",
        "confidence": 0.0,
        "text_region": [[5.0, 59.0], [23.0, 59.0], [23.0, 77.0], [5.0, 77.0]],
    },
    {
        "text": "2",
        "confidence": 0.9991047978401184,
        "text_region": [[83.0, 60.0], [96.0, 60.0], [96.0, 74.0], [83.0, 74.0]],
    },
    {
        "text": "140,986.93",
        "confidence": 0.9872023463249207,
        "text_region": [[177.0, 59.0], [262.0, 59.0], [262.0, 77.0], [177.0, 77.0]],
    },
    {
        "text": "14,000EA",
        "confidence": 0.9725096225738525,
        "text_region": [[385.0, 59.0], [457.0, 59.0], [457.0, 77.0], [385.0, 77.0]],
    },
    {
        "text": "器",
        "confidence": 0.16805972158908844,
        "text_region": [[483.0, 61.0], [497.0, 61.0], [497.0, 73.0], [483.0, 73.0]],
    },
    {
        "text": "4200001437",
        "confidence": 0.9967278242111206,
        "text_region": [[506.0, 59.0], [592.0, 59.0], [592.0, 77.0], [506.0, 77.0]],
    },
    {
        "text": "140",
        "confidence": 0.9974072575569153,
        "text_region": [[612.0, 57.0], [644.0, 57.0], [644.0, 77.0], [612.0, 77.0]],
    },
    {
        "text": "SocialWelfareSurch",
        "confidence": 0.9858036041259766,
        "text_region": [[670.0, 59.0], [803.0, 59.0], [803.0, 77.0], [670.0, 77.0]],
    },
    {
        "text": "GI（GST",
        "confidence": 0.9061947464942932,
        "text_region": [[1073.0, 55.0], [1164.0, 58.0], [1164.0, 79.0], [1072.0, 76.0]],
    },
    {
        "text": "鑫",
        "confidence": 0.19185461103916168,
        "text_region": [[1173.0, 61.0], [1186.0, 61.0], [1186.0, 74.0], [1173.0, 74.0]],
    },
    {
        "text": "?",
        "confidence": 0.28549399971961975,
        "text_region": [[5.0, 84.0], [23.0, 84.0], [23.0, 102.0], [5.0, 102.0]],
    },
    {
        "text": "3",
        "confidence": 0.9981780052185059,
        "text_region": [[84.0, 84.0], [96.0, 84.0], [96.0, 100.0], [84.0, 100.0]],
    },
    {
        "text": "604,229.70",
        "confidence": 0.971087634563446,
        "text_region": [[176.0, 82.0], [261.0, 82.0], [261.0, 100.0], [176.0, 100.0]],
    },
    {
        "text": "6,000EA",
        "confidence": 0.9679980278015137,
        "text_region": [[391.0, 82.0], [459.0, 82.0], [459.0, 103.0], [391.0, 103.0]],
    },
    {
        "text": "口",
        "confidence": 0.9026997089385986,
        "text_region": [[480.0, 83.0], [510.0, 83.0], [510.0, 102.0], [480.0, 102.0]],
    },
    {
        "text": "4200001437",
        "confidence": 0.9955192804336548,
        "text_region": [[502.0, 84.0], [592.0, 84.0], [592.0, 102.0], [502.0, 102.0]],
    },
    {
        "text": "150",
        "confidence": 0.9983319640159607,
        "text_region": [[612.0, 83.0], [643.0, 83.0], [643.0, 102.0], [612.0, 102.0]],
    },
    {
        "text": "INBasiccustoms",
        "confidence": 0.989190399646759,
        "text_region": [[670.0, 84.0], [777.0, 84.0], [777.0, 102.0], [670.0, 102.0]],
    },
    {
        "text": "GI(GST",
        "confidence": 0.9113300442695618,
        "text_region": [
            [1074.0, 82.0],
            [1167.0, 82.0],
            [1167.0, 103.0],
            [1074.0, 103.0],
        ],
    },
    {
        "text": "口",
        "confidence": 0.9762108325958252,
        "text_region": [
            [1169.0, 84.0],
            [1187.0, 84.0],
            [1187.0, 101.0],
            [1169.0, 101.0],
        ],
    },
    {
        "text": "√",
        "confidence": 0.35394710302352905,
        "text_region": [[5.0, 109.0], [23.0, 109.0], [23.0, 123.0], [5.0, 123.0]],
    },
    {
        "text": "4",
        "confidence": 0.9983332753181458,
        "text_region": [[84.0, 110.0], [96.0, 110.0], [96.0, 121.0], [84.0, 121.0]],
    },
    {
        "text": "60,422.97",
        "confidence": 0.9447630643844604,
        "text_region": [[186.0, 110.0], [261.0, 110.0], [261.0, 123.0], [186.0, 123.0]],
    },
    {
        "text": "6.000EA",
        "confidence": 0.9239044785499573,
        "text_region": [[394.0, 110.0], [456.0, 110.0], [456.0, 123.0], [394.0, 123.0]],
    },
    {
        "text": "国",
        "confidence": 0.11650555580854416,
        "text_region": [[483.0, 110.0], [497.0, 110.0], [497.0, 120.0], [483.0, 120.0]],
    },
    {
        "text": "4200001437",
        "confidence": 0.9987810254096985,
        "text_region": [[507.0, 110.0], [591.0, 110.0], [591.0, 123.0], [507.0, 123.0]],
    },
    {
        "text": "150",
        "confidence": 0.9992054104804993,
        "text_region": [[613.0, 109.0], [641.0, 109.0], [641.0, 123.0], [613.0, 123.0]],
    },
    {
        "text": "Social WelfareSurch",
        "confidence": 0.9671373963356018,
        "text_region": [[670.0, 109.0], [801.0, 109.0], [801.0, 123.0], [670.0, 123.0]],
    },
    {
        "text": "GI（GST.",
        "confidence": 0.919130802154541,
        "text_region": [
            [1076.0, 110.0],
            [1141.0, 110.0],
            [1141.0, 123.0],
            [1076.0, 123.0],
        ],
    },
    {
        "text": "圈",
        "confidence": 0.13210679590702057,
        "text_region": [
            [1173.0, 110.0],
            [1186.0, 110.0],
            [1186.0, 121.0],
            [1173.0, 121.0],
        ],
    },
]

def make_excel(result, output_path:str = '.', filename:str = 'output.xlsx'):
# Create a new Excel workbook
    import os
    import pandas as pd
    import numpy as np
    workbook = openpyxl.Workbook()
    worksheet = workbook.active

    # Add headers for the columns
    worksheet.cell(row=1, column=1, value="Text")
    worksheet.cell(row=1, column=2, value="Confidence")
    worksheet.cell(row=1, column=3, value="x_min")
    worksheet.cell(row=1, column=4, value="y_min")
    worksheet.cell(row=1, column=5, value="x_max")
    worksheet.cell(row=1, column=6, value="y_max")

    # Iterate through the dictionary and insert data into the worksheet
    for idx, data in enumerate(result, start=2):
        text = data["text"]
        confidence = data["confidence"]
        text_region = data["text_region"]

        x_min = min(point[0] for point in text_region)
        y_min = min(point[1] for point in text_region)
        x_max = max(point[0] for point in text_region)
        y_max = max(point[1] for point in text_region)

        worksheet.cell(row=idx, column=1, value=text)
        worksheet.cell(row=idx, column=2, value=confidence)
        worksheet.cell(row=idx, column=3, value=x_min)
        worksheet.cell(row=idx, column=4, value=y_min)
        worksheet.cell(row=idx, column=5, value=x_max)
        worksheet.cell(row=idx, column=6, value=y_max)

    df = pd.DataFrame(workbook.active.values)
    df.columns = df.iloc[0]
    differences = np.diff(df['x_max'])
    indices = np.where(-differences > 500)[0]
    print(df.iloc[indices,:])
    df['Text'].iloc[indices+1]
    df['Text'].iloc[indices]
    data = df['x_max'].values
    subarrays = np.split(data, indices + 1)
    new_df = pd.DataFrame(subarrays)
    new_df.to_excel(output_path, filename.split('.')[0]+'s.xlsx')
    new_df.to_html(os.path.join(output_path, filename.split('.')[0]+'s.html'))

if __name__ == '__main__':
    make_excel(result)